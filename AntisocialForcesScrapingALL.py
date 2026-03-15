
"""`
   Excelから検索Wordを取得し、それを上から順にGoogleで検索をかけ、
   その一番上に出てきたタイトルと、そのページのtitleタグ中の要素をExcelに書き出し保存する。（スクレイピング）

  採用技術
   言語 Python (HTML)
    技術 スクレイピング クローリング   

  使用上の注意 
   ・プログラムを起動する際にはExcelを開かないようにしてください。
   Excelを開いた状態で起動させると、熱暴走したり、インプットするExcelが壊れてしまうことがあります。

   11/8 検索結果が10件未満の時に次の会社の出力が狂うのを直したい
   11/10検索結果がきれいになった。検証資料を作成し、和司さんに見せる。
      また、赤字になった箇所を数える機能を追加する。
      動画がヒットした時だけ、出力がおかしくなる。（検索ワードそのまま＋「の動画」と出力される）
      見かけが悪いが、反社チェックという意味では問題無いと思われる。
   11/18コードがスパゲッティになってきたのでリファクタリングしたい。
       クエリパラメータを削除するコードを追加したい。
   
   11/24大量のデータを流した結果、何故か途中から最初の会社に戻る現象が生じた。
      原因の究明をするか、或いは少量のデータのみで使用してもらうことにするか、、、。

   12/03どうも90件くらいのところで過重負荷なのか、処理が停止してしまう。

   12/10大体50件刻みでやると効率的である。これ以上だと熱暴走（？）してとまってしまう。

   12/21現時点で認識しているバグを一通り潰せた。
           サイバー攻撃を疑われないようにするためである。

      23/7/25 インプットデータに新しい列が新設される「申請者」「担当者」ので、Excelへの転記の仕方を調節する
   　　　　　また、使用しないコメントアウト行を削除した
   
"""

import requests
from requests.exceptions import ConnectionError, TooManyRedirects, HTTPError
from bs4 import BeautifulSoup
import bs4
import time
import openpyxl
import subprocess
import random
import time
import re

ANTISOCIAL＿FORCES = ' AND (○○ OR △△ OR □□ OR ☆☆ OR '\
'○○ OR △△ OR □□ OR ☆☆ OR '\)

ANTISOCIAL_FORCES_LIST = ["○○" , "△△" , "□□" , "☆☆"\
    "○○" , "△△" , "□□" , "☆☆"\]

#相手サーバに負担をかけないために、タイムスリープを設定（１秒以上取っておくと安全）
#SLEEP_TIME = 1

#2列目以降に検索結果を次々と書き込んでいく
r = 5
# 時間計測開始
time_sta = time.time()
# 上位から何件までのサイトを抽出するか指定する
RANK = 20
pages_num = RANK + 1

AntisocialCount = 0

inputCellRow = 2
inputCellcolumn = 5

titleCounter = 0

writtenCellRow = 2
writtenCellcolumn = 6

#インプットデータ（Excel）を配置している絶対パスを設定する
wb = openpyxl.load_workbook( r'C:\Users' )

#シート取得
ws = wb[ "Sheet1" ]

#シートの行数を取得する。
maxRow = wb['Sheet1'].max_row

# 背景色を設定
fill = openpyxl.styles.PatternFill(patternType='solid',
                                   fgColor='ffd700', bgColor='ffd700')

for i in range( 2 , maxRow + 1 ):
    #Excelに入力された業者or顧客名を取得する
    
    if titleCounter == 0:
        cell = ws.cell( row = inputCellRow , column = inputCellcolumn )
    
    #見出し語20個目なら行だけインクリメントしてインプット（代表者のインプット）
    if titleCounter == RANK:
        cell = ws.cell( row = inputCellRow , column = inputCellcolumn + 1 )
    
    #見出し語40個目なら列をインクリメントして、インプット（次の会社名をインプット） 見出し語カウンターを0に戻す
    if titleCounter >=  RANK * 2:
        titleCounter = 0
        inputCellRow += 1
        writtenCellcolumn = 6
        cell = ws.cell( row = inputCellRow , column = inputCellcolumn )

    #検索ワードがなくなったら終了する
    if cell.value is None:
        print("検索ワードが無いので終了！")
        break
    
    #上で取得した業者or顧客名をstring型にキャストする。
    cellString = str( cell.value )

    #「㈱」と「㈲」を削除する（新聞記事は会社名を基本的に「㈱」と「㈲」を抜いて記載しており、新聞の記事を検索にヒットさせるため
    # 「医療法人」「株式会社」「(一社)」「合同会社」「土地家屋調査士法人」「(同)」「一般財団法人」「(有)」「合資会社」「土質試験協同組合」

    cellString_replaced1 = cellString.replace('㈱', '')
    cellString_replaced2 = cellString_replaced1.replace('㈲', '')
    cellString_replaced3 = cellString_replaced2.replace('医療法人', '')
    cellString_replaced4 = cellString_replaced3.replace('株式会社', '')
    cellString_replaced5 = cellString_replaced4.replace('(一社)', '')
    cellString_replaced6 = cellString_replaced5.replace('合同会社', '')
    cellString_replaced7 = cellString_replaced6.replace('土地家屋調査士法人', '')
    cellString_replaced8 = cellString_replaced7.replace('(同)', '')
    cellString_replaced9 = cellString_replaced8.replace('一般財団法人', '')
    cellString_replaced10 = cellString_replaced9.replace('(有)', '')
    cellString_replaced11 = cellString_replaced10.replace('合資会社', '')
    cellString_replaced12 = cellString_replaced11.replace('土質試験協同組合', '')

    # 「代表取締役」「院長」「総経理」「学長」「代表理事」「President」「最高経営責任者」「弁護士」「取締役社長」「Director」「所長」
    cellString_replaced13 = cellString_replaced12.replace('代表取締役', '')
    cellString_replaced14 = cellString_replaced13.replace('院長', '')
    cellString_replaced15 = cellString_replaced14.replace('総経理', '')
    cellString_replaced16 = cellString_replaced15.replace('学長', '')
    cellString_replaced17 = cellString_replaced16.replace('代表理事', '')
    cellString_replaced18 = cellString_replaced17.replace('President', '')
    cellString_replaced19 = cellString_replaced18.replace('最高経営責任者', '')
    cellString_replaced20 = cellString_replaced19.replace('弁護士', '')
    cellString_replaced21 = cellString_replaced20.replace('取締役社長', '')
    cellString_replaced22 = cellString_replaced21.replace('Director', '')
    cellString_replaced23 = cellString_replaced22.replace('所長', '')

    # 「代表社員」「工場長」「取締役」「支店長」「代表」「共同代表」「次長」「代表執行役」「代表者」「会長」
    cellString_replaced24 = cellString_replaced23.replace('代表社員', '')
    cellString_replaced25 = cellString_replaced24.replace('工場長', '')
    cellString_replaced26 = cellString_replaced25.replace('取締役', '')
    cellString_replaced27 = cellString_replaced26.replace('支店長', '')
    cellString_replaced28 = cellString_replaced27.replace('共同代表', '')
    cellString_replaced29 = cellString_replaced28.replace('次長', '')
    cellString_replaced30 = cellString_replaced29.replace('代表執行役', '')
    cellString_replaced31 = cellString_replaced30.replace('代表者', '')
    cellString_replaced32 = cellString_replaced31.replace('代表', '')
    cellString_replaced33 = cellString_replaced32.replace('会長', '')
    cellString_replaced34 = cellString_replaced33.replace('社長', '')
    cellString_replaced35 = cellString_replaced34.replace('CEO', '')
    cellString_replaced36 = cellString_replaced35.replace('/', '')
    cellString_replaced37 = cellString_replaced36.replace('兼', '')
    cellString_replaced38 = cellString_replaced37.replace('Partner', '')
    cellString_replaced39 = cellString_replaced38.replace('　', '')
    cellString_replaced40 = cellString_replaced39.replace(' ', '')
    
    # テスト用出力
    #ws.cell( row = i, column = 2  ).value = cellString_replaced40
    
    #業者or顧客名に反社ワードを連結させる。
    search_word = cellString_replaced40 + ANTISOCIAL＿FORCES 
    
    #ここを通る時、rは2に戻す。
    r = 5
    
    print( f'【検索ワード】{search_word}' )
    
    # Googleから検索結果ページを取得する
    url = f'https://www.google.co.jp/search?hl=ja&num={pages_num}&q={search_word}'
    request = requests.get( url )
    print(request)
    # Googleのページ解析を行う
    soup = BeautifulSoup( request.text, "html.parser" )
    search_site_list = soup.select( 'div.kCrYT > a' )

    # ページ解析と結果の出力
    for rank, site in zip(range( 1 , pages_num ) , search_site_list):
            #見出し語のカウント
            titleCounter += 1
            
            try:
                #site.select('h3.zBAuLc')[0]では、<h3 class="zBAuLc"/>サイトタイトル/</h3>部分を抽出しています。
                site_title = site.select( 'h3' )[0].text
            except IndexError:
                site_title = site.select( 'img' )[0]['alt']
                continue
            site_url = site['href'].replace( '/url?q=', '' )
            
            #余分な文字列（クエリパラメータ）を削除する。
            if '&' in site_url:
                print('URLから&を検出')
                txt = site_url
                anp = txt.find('&')
                site_url = txt[:anp] 
                print( 'site_url = ' + site_url )    

            if '%' in site_url:
                print('URLから%を検出')
                per = site_url.find('%')
                site_url = site_url[:per]     
                print( 'site_url = ' + site_url ) 
            
            if '?' in site_url:
                print('URLから?を検出')
                qes = site_url.find('?')
                site_url = site_url[:qes]     
                print( 'site_url = ' + site_url ) 

            print('最終的なURL = ' + site_url)
            
            # 結果を出力する
            print( str(rank) + "位: " + site_title )
            print( "!!!!!!!!!!!!!!!!!!!!!! titleCounter = " + str(titleCounter)  + " !!!!!!!!!!!!!!!!!!!!!! ")
            
            writtenCellcolumn += 1
            ws.cell( row = inputCellRow, column = writtenCellcolumn ).value = site_title
            ws.cell( row = inputCellRow, column = writtenCellcolumn ).hyperlink = site_url
    
            #ユーザエージェントを設定　「私は○○と言うブラウザです！」と自己紹介するためのコード。これが無いと止まる。
            try:
                headers_dic = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36"}
                res = requests.get( site_url, headers=headers_dic )
                
            #例外処理　プログラム内で対応可能なエラーをスルーする
            except  TooManyRedirects as e:
                print('catch TooManyRedirects:', e)
                r = r + 1
                if r == RANK + r:
                    r = 5
                continue
            except  ConnectionError as e:
                print('catch ConnectionError:', e)
                r = r + 1
                if r == RANK + r:
                    r = 5
                continue

            #取得した見出し語について、反社ワードを１つ１つ照合して、反社ワードが含まれる見出し語が書き込まれたセルをハイライトする。
            for x in range(0,127):
                if ANTISOCIAL_FORCES_LIST[x] in site_title: 
                    print('!!!!!!!!!!反社ワード走査!!!!!!!!!!!!' + ANTISOCIAL_FORCES_LIST[x])
                    ws.cell( row = inputCellRow, column = writtenCellcolumn ).fill = fill

            #インクリメントと調整
            r = r + 1
            if r == RANK + r:
                r = 5
            
            #ハッキングやサイバー攻撃（DOS攻撃）を疑われないためにあえて処理を遅く調節している。  
            #相手サーバに負担をかけないために、タイムスリープを設定（１秒以上取っておくと安全）            
            #一回ごとに保存した方がリスクが小さいかも知れない（動作が遅くなったりするかも？）
            wb.save( r'C:\Users\○○\Desktop\2024年度反社チェック\○○.xlsx' )
            SLEEP_TIME = random.randrange(2)

            time.sleep( SLEEP_TIME )


wb.save( r'C:\Users\○○\Desktop\2024年度反社チェック\○○.xlsx' )
# 時間計測終了
time_end = time.time()
# 経過時間（秒）
tim = time_end- time_sta

print( str( tim ) + "秒かかりました。")
wb.close()
